import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import plotly.express as px

# PAGE CONFIG
st.set_page_config(
    page_title="Customer Service",
    layout="wide"
)

st.title("Infra vs Universe Cases")

# LOAD EXCEL
df = pd.read_excel("sn_customerservice_case.xlsx")

# CLEAN COLUMN NAMES
df.columns = df.columns.str.strip().str.lower()

# LOAD WORKBOOK
wb = load_workbook("sn_customerservice_case.xlsx")
ws = wb["Infra EE4 Cases"]  # Change if sheet name is different

infra_count = 0
universe_count = 0

# Short Description column = D = index 3
for row in ws.iter_rows(min_row=2):
    cell = row[3]

    color = cell.fill.fgColor.rgb

    if color == "FF00FF00":      # Green
        infra_count += 1

    elif color == "FFFF0000":    # Red
        universe_count += 1

# Create DataFrame for Pie Chart
team_summary = pd.DataFrame({
    "Team": ["Infra", "Universe"],
    "Count": [infra_count, universe_count]
})

# Show counts
col1, col2 = st.columns(2)

with col1:
    st.metric("Infra Cases", infra_count)

with col2:
    st.metric("Universe Cases", universe_count)

# Pie Chart
fig = px.pie(
    team_summary,
    names="Team",
    values="Count",
    title="Infra vs Universe Distribution"
)

st.plotly_chart(fig, use_container_width=True)
